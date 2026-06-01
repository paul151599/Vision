"""Baseline creation, loading, saving, and Excel import."""
import json
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

from config.settings import BASELINE_DIR
from config.camera import CameraConfig
from config.feature_registry import FEATURE_REGISTRY
from core.data_models import BaselineStats, BaselineSpec, ProcessType
from core.feature_engine import FeatureEngine


class BaselineManager:
    """Create and manage baseline specifications."""

    def __init__(self, camera_config: Optional[CameraConfig] = None):
        self.engine = FeatureEngine(camera_config)
        self.camera_config = camera_config or CameraConfig()
        BASELINE_DIR.mkdir(parents=True, exist_ok=True)

    def create_from_images(
        self,
        image_paths: List[str | Path],
        sample_id: str,
        process_type: ProcessType,
    ) -> BaselineSpec:
        """Create baseline from a set of reference images."""
        all_values: Dict[str, List[float]] = {}

        for path in image_paths:
            features = self.engine.extract_values(path)
            for fid, val in features.items():
                all_values.setdefault(fid, []).append(val)

        stats = {}
        for fid, vals in all_values.items():
            arr = np.array(vals)
            stats[fid] = BaselineStats(
                feature_id=fid,
                mean=float(np.mean(arr)),
                std=float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0,
                min=float(np.min(arr)),
                max=float(np.max(arr)),
                n_samples=len(arr),
                values=vals,
            )

        return BaselineSpec(
            process_type=process_type,
            sample_id=sample_id,
            camera_config=self.camera_config.to_dict(),
            stats=stats,
            n_samples=len(image_paths),
        )

    def save(self, baseline: BaselineSpec, path: Optional[Path] = None) -> Path:
        """Save baseline to JSON."""
        if path is None:
            fname = f"{baseline.sample_id}_{baseline.process_type.value}.json"
            path = BASELINE_DIR / fname

        data = {
            "process_type": baseline.process_type.value,
            "sample_id": baseline.sample_id,
            "camera_config": baseline.camera_config,
            "created_at": baseline.created_at,
            "n_samples": baseline.n_samples,
            "stats": {
                fid: {
                    "mean": s.mean,
                    "std": s.std,
                    "min": s.min,
                    "max": s.max,
                    "n_samples": s.n_samples,
                    "values": s.values,
                }
                for fid, s in baseline.stats.items()
            },
        }
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        return path

    def load(self, path: str | Path) -> BaselineSpec:
        """Load baseline from JSON."""
        path = Path(path)
        data = json.loads(path.read_text(encoding="utf-8"))

        stats = {}
        for fid, s in data["stats"].items():
            stats[fid] = BaselineStats(
                feature_id=fid,
                mean=s["mean"],
                std=s["std"],
                min=s["min"],
                max=s["max"],
                n_samples=s["n_samples"],
                values=s.get("values", []),
            )

        return BaselineSpec(
            process_type=ProcessType(data["process_type"]),
            sample_id=data["sample_id"],
            camera_config=data.get("camera_config", {}),
            stats=stats,
            created_at=data.get("created_at", ""),
            n_samples=data.get("n_samples", 0),
        )

    def import_from_excel(
        self,
        excel_path: str | Path,
        sample_id: str,
        process_type: ProcessType,
    ) -> BaselineSpec:
        """Import baseline from pellicle_baseline_spec Excel file.

        Expects the Excel to have feature values for Pt1~Pt5 and
        computed Mean/Std columns.
        """
        import openpyxl

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # Find header row and column mapping
        headers = {}
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "feature":
                    header_row = cell.row
                    for hcell in ws[header_row]:
                        if hcell.value:
                            headers[str(hcell.value).strip()] = hcell.column
                    break
            if header_row:
                break

        if not header_row:
            raise ValueError("Could not find header row with 'Feature' column in Excel")

        # Find Pt columns and Mean/Std columns
        pt_cols = []
        mean_col = None
        std_col = None
        for name, col in headers.items():
            lower = name.lower()
            if lower.startswith("pt") or lower.startswith("point") or lower.startswith("#"):
                pt_cols.append(col)
            elif lower in ("mean", "avg", "average"):
                mean_col = col
            elif lower in ("std", "stdev", "sd"):
                std_col = col

        feature_col = headers.get("Feature") or headers.get("feature") or headers.get("Feature ID")

        stats = {}
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            fid = None
            for cell in row:
                if cell.column == feature_col:
                    fid = str(cell.value).strip() if cell.value else None
                    break
            if not fid:
                continue

            # Collect point values
            values = []
            for pc in pt_cols:
                val = row[pc - 1].value
                if val is not None:
                    try:
                        values.append(float(val))
                    except (ValueError, TypeError):
                        pass

            if not values:
                continue

            arr = np.array(values)
            stats[fid] = BaselineStats(
                feature_id=fid,
                mean=float(np.mean(arr)),
                std=float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0,
                min=float(np.min(arr)),
                max=float(np.max(arr)),
                n_samples=len(arr),
                values=values,
            )

        return BaselineSpec(
            process_type=process_type,
            sample_id=sample_id,
            stats=stats,
            n_samples=len(pt_cols),
        )
