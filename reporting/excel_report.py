"""Excel report generation compatible with v3.1 format."""
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR
from config.feature_registry import FEATURE_REGISTRY, get_feature
from core.data_models import InspectionResult, Judgment, BaselineSpec


# Style constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NA_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_report(
    results: List[InspectionResult],
    baseline: Optional[BaselineSpec] = None,
    output_path: Optional[Path] = None,
) -> Path:
    """Generate Excel inspection report.

    Args:
        results: List of inspection results
        baseline: Optional baseline for spec columns
        output_path: Output file path (default: output/inspection_report.xlsx)

    Returns:
        Path to generated report
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        output_path = OUTPUT_DIR / "inspection_report.xlsx"

    wb = openpyxl.Workbook()

    # -- Summary sheet --
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, results)

    # -- Detail sheet --
    ws_detail = wb.create_sheet("Feature Details")
    _write_detail(ws_detail, results, baseline)

    wb.save(str(output_path))
    return output_path


def _write_summary(ws, results: List[InspectionResult]):
    """Write summary sheet with overall results."""
    headers = ["Image", "Process", "Overall", "Fail Count", "Critical Fails"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, res in enumerate(results, 2):
        ws.cell(row=i, column=1, value=Path(res.image_path).name).border = THIN_BORDER
        ws.cell(row=i, column=2, value=res.process_type.value).border = THIN_BORDER

        overall_cell = ws.cell(row=i, column=3, value=res.overall_judgment.value)
        overall_cell.fill = PASS_FILL if res.overall_judgment == Judgment.PASS else FAIL_FILL
        overall_cell.border = THIN_BORDER
        overall_cell.alignment = Alignment(horizontal="center")

        fail_count = sum(1 for j in res.judgments.values() if j.judgment == Judgment.FAIL)
        ws.cell(row=i, column=4, value=fail_count).border = THIN_BORDER

        critical = [r for r in res.fail_reasons if "[CRITICAL]" in r]
        ws.cell(row=i, column=5, value=len(critical)).border = THIN_BORDER

    # Auto-width
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _write_detail(ws, results: List[InspectionResult], baseline: Optional[BaselineSpec]):
    """Write detailed feature values and judgments."""
    # Headers
    headers = ["Feature ID", "Category", "Group"]
    if baseline:
        headers += ["Baseline Mean", "Baseline Std", "Lower Spec", "Upper Spec"]
    for res in results:
        name = Path(res.image_path).stem
        headers += [f"{name} Value", f"{name} Judge"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Feature rows
    for row_idx, fspec in enumerate(FEATURE_REGISTRY, 2):
        col = 1
        ws.cell(row=row_idx, column=col, value=fspec.id).border = THIN_BORDER
        col += 1
        ws.cell(row=row_idx, column=col, value=fspec.category).border = THIN_BORDER
        col += 1
        ws.cell(row=row_idx, column=col, value=fspec.group.value).border = THIN_BORDER
        col += 1

        if baseline and fspec.id in baseline.stats:
            st = baseline.stats[fspec.id]
            ws.cell(row=row_idx, column=col, value=round(st.mean, 6)).border = THIN_BORDER
            col += 1
            ws.cell(row=row_idx, column=col, value=round(st.std, 6)).border = THIN_BORDER
            col += 1
            # Spec bounds from first result's judgment
            if results and fspec.id in results[0].judgments:
                jdg = results[0].judgments[fspec.id]
                ws.cell(row=row_idx, column=col, value=round(jdg.lower_spec, 6) if jdg.lower_spec is not None else "N/A").border = THIN_BORDER
                col += 1
                ws.cell(row=row_idx, column=col, value=round(jdg.upper_spec, 6) if jdg.upper_spec is not None else "N/A").border = THIN_BORDER
                col += 1
            else:
                col += 2
        elif baseline:
            col += 4

        for res in results:
            if fspec.id in res.features:
                val = res.features[fspec.id].value
                ws.cell(row=row_idx, column=col, value=round(val, 6)).border = THIN_BORDER
            col += 1

            if fspec.id in res.judgments:
                jdg = res.judgments[fspec.id]
                jcell = ws.cell(row=row_idx, column=col, value=jdg.judgment.value)
                jcell.border = THIN_BORDER
                jcell.alignment = Alignment(horizontal="center")
                if jdg.judgment == Judgment.PASS:
                    jcell.fill = PASS_FILL
                elif jdg.judgment == Judgment.FAIL:
                    jcell.fill = FAIL_FILL
                else:
                    jcell.fill = NA_FILL
            col += 1

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
