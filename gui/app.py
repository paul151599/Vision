"""EUV Pellicle Vision Inspection System - Feature Extraction & Analysis Tool."""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import threading
import sys
import csv

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def _image_to_clipboard(pil_img):
    """Place a PIL image on the Windows clipboard as CF_DIB.

    CF_DIB is the format Excel / PowerPoint accept on paste. Implemented with
    ctypes (no pywin32 dependency). A BMP file is just a 14-byte
    BITMAPFILEHEADER followed by the DIB, so we strip those 14 bytes.
    """
    import io
    import ctypes
    from ctypes import c_void_p, c_size_t, c_uint, memmove

    rgb = pil_img.convert("RGB")
    buf = io.BytesIO()
    rgb.save(buf, "BMP")
    data = buf.getvalue()[14:]  # strip BITMAPFILEHEADER → CF_DIB payload
    buf.close()

    CF_DIB = 8
    GMEM_MOVEABLE = 0x0002

    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    # Set pointer-sized return/arg types so 64-bit handles are not truncated.
    kernel32.GlobalAlloc.restype = c_void_p
    kernel32.GlobalAlloc.argtypes = [c_uint, c_size_t]
    kernel32.GlobalLock.restype = c_void_p
    kernel32.GlobalLock.argtypes = [c_void_p]
    kernel32.GlobalUnlock.argtypes = [c_void_p]
    user32.OpenClipboard.argtypes = [c_void_p]
    user32.SetClipboardData.restype = c_void_p
    user32.SetClipboardData.argtypes = [c_uint, c_void_p]

    if not user32.OpenClipboard(None):
        raise OSError("클립보드를 열 수 없습니다.")
    try:
        user32.EmptyClipboard()
        h_global = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(data))
        if not h_global:
            raise MemoryError("GlobalAlloc 실패")
        p = kernel32.GlobalLock(h_global)
        memmove(p, data, len(data))
        kernel32.GlobalUnlock(h_global)
        if not user32.SetClipboardData(CF_DIB, h_global):
            raise OSError("SetClipboardData 실패")
        # On success the clipboard owns the memory; do not free h_global.
    finally:
        user32.CloseClipboard()


def _save_image_lossless(pil_img, path):
    """Save a PIL image without lossy compression.

    PNG is stored with compress_level=0 (no compression), BMP/TIFF are written
    uncompressed. Format is chosen from the file extension.
    """
    ext = Path(path).suffix.lower()
    if ext == ".png":
        pil_img.save(path, "PNG", compress_level=0)
    elif ext in (".bmp", ".dib"):
        pil_img.convert("RGB").save(path, "BMP")
    elif ext in (".tif", ".tiff"):
        pil_img.save(path, "TIFF", compression="none")
    else:
        pil_img.save(path)


class InspectionApp(tk.Tk):
    """Feature extraction and visualization tool."""

    def __init__(self):
        super().__init__()
        self.title("Feature Extraction Tool v1.0")
        self.geometry("1600x900")
        self.minsize(1200, 700)

        # State
        self.image_paths = []
        self.current_index = -1
        self.feature_values = {}   # path -> Dict[str, float]
        self.phase_images = {}     # path -> Dict[phase_name, PIL.Image]
        self.processing = False
        self.proc_times = {}       # path -> seconds

        # Style
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Header.TLabel", font=("Segoe UI", 12, "bold"))
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=22)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))

        self._build_menu()
        self._build_layout()
        self._bind_events()

    def _build_menu(self):
        menubar = tk.Menu(self)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open Image(s)...", command=self._open_files, accelerator="Ctrl+O")
        file_menu.add_command(label="Open Folder...", command=self._open_folder, accelerator="Ctrl+Shift+O")
        file_menu.add_separator()
        file_menu.add_command(label="Export CSV...", command=self._export_csv)
        file_menu.add_command(label="Export Excel...", command=self._export_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Export All Images (전체 이미지 다운로드)...",
                              command=self._export_all_images)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)
        menubar.add_cascade(label="File", menu=file_menu)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Capture Conditions (촬영 조건 Ref)...",
                              command=self._show_capture_conditions)
        view_menu.add_separator()
        view_menu.add_command(label="Clear All", command=self._clear_all)
        menubar.add_cascade(label="View", menu=view_menu)

        self.config(menu=menubar)

    def _build_layout(self):
        self.main_pane = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.main_pane.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ── Left Panel: File list + controls ──
        left_frame = ttk.Frame(self.main_pane, width=250)
        self.main_pane.add(left_frame, weight=0)

        ttk.Label(left_frame, text="Images", style="Header.TLabel").pack(pady=(5, 2), padx=5, anchor="w")

        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Button(btn_frame, text="+ Files", width=8, command=self._open_files).pack(side=tk.LEFT, padx=1)
        ttk.Button(btn_frame, text="+ Folder", width=8, command=self._open_folder).pack(side=tk.LEFT, padx=1)
        ttk.Button(btn_frame, text="Clear", width=6, command=self._clear_all).pack(side=tk.LEFT, padx=1)

        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)
        self.image_listbox = tk.Listbox(list_frame, font=("Segoe UI", 9), selectmode=tk.EXTENDED)
        list_scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.image_listbox.yview)
        self.image_listbox.config(yscrollcommand=list_scroll.set)
        self.image_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Run button + progress
        ctrl_frame = ttk.LabelFrame(left_frame, text="Analysis")
        ctrl_frame.pack(fill=tk.X, padx=5, pady=5)

        self.run_btn = ttk.Button(ctrl_frame, text="▶  Extract Features", command=self._run_analysis)
        self.run_btn.pack(fill=tk.X, padx=5, pady=5)

        self.progress = ttk.Progressbar(ctrl_frame, mode="determinate")
        self.progress.pack(fill=tk.X, padx=5, pady=(0, 5))
        self.status_label = ttk.Label(ctrl_frame, text="Ready", foreground="gray")
        self.status_label.pack(padx=5, pady=(0, 5))

        # ── Center Panel: Phase visualization tabs ──
        center_frame = ttk.Frame(self.main_pane)
        self.main_pane.add(center_frame, weight=3)

        self.view_notebook = ttk.Notebook(center_frame)
        self.view_notebook.pack(fill=tk.BOTH, expand=True)

        self.phase_tabs = {}
        self.phase_canvases = {}
        self.phase_photo_refs = {}

        tab_names = [
            ("Original", "Original"),
            ("Grayscale", "Grayscale"),
            ("Channels", "R/G/B Channels"),
            ("GainCorrected", "Gain Corrected"),
            ("FFT", "FFT Spectrum"),
            ("Anomaly", "Anomaly Map"),
            ("LocalContrast", "Local Contrast"),
            ("Entropy", "Entropy Map"),
            ("CIELAB", "CIELAB dE"),
            ("Morphology", "Morphology"),
        ]

        for tab_id, tab_label in tab_names:
            frame = ttk.Frame(self.view_notebook)
            self.view_notebook.add(frame, text=tab_label)

            # Header bar: phase title (left) + action buttons (top-right)
            header = ttk.Frame(frame)
            header.pack(fill=tk.X, side=tk.TOP)
            ttk.Label(header, text=tab_label, style="Header.TLabel").pack(
                side=tk.LEFT, padx=8, pady=3)
            # Packed right-to-left → visual order: [복사] [다운로드] [ⓘ 처리 설명]
            ttk.Button(
                header, text="ⓘ  처리 설명", width=14,
                command=lambda tid=tab_id: self._show_phase_info(tid),
            ).pack(side=tk.RIGHT, padx=(2, 8), pady=3)
            ttk.Button(
                header, text="다운로드", width=10,
                command=lambda tid=tab_id: self._download_phase_image(tid),
            ).pack(side=tk.RIGHT, padx=2, pady=3)
            ttk.Button(
                header, text="복사", width=8,
                command=lambda tid=tab_id: self._copy_phase_image(tid),
            ).pack(side=tk.RIGHT, padx=2, pady=3)

            canvas = tk.Canvas(frame, bg="#2b2b2b")
            canvas.pack(fill=tk.BOTH, expand=True)
            self.phase_tabs[tab_id] = frame
            self.phase_canvases[tab_id] = canvas
            self.phase_photo_refs[tab_id] = None

        # ── Right Panel: Feature values + Comparison ──
        right_frame = ttk.Frame(self.main_pane, width=420)
        self.main_pane.add(right_frame, weight=1)

        right_notebook = ttk.Notebook(right_frame)
        right_notebook.pack(fill=tk.BOTH, expand=True)

        # Tab 1: Feature values for selected image
        feat_frame = ttk.Frame(right_notebook)
        right_notebook.add(feat_frame, text="Feature Values")

        self.feat_tree = ttk.Treeview(
            feat_frame,
            columns=("feature", "value", "info"),
            show="headings",
            selectmode="browse",
        )
        self.feat_tree.heading("feature", text="Feature")
        self.feat_tree.heading("value", text="Value")
        self.feat_tree.heading("info", text="")
        self.feat_tree.column("feature", width=200)
        self.feat_tree.column("value", width=120, anchor="e")
        self.feat_tree.column("info", width=30, anchor="center")

        feat_scroll = ttk.Scrollbar(feat_frame, orient=tk.VERTICAL, command=self.feat_tree.yview)
        self.feat_tree.config(yscrollcommand=feat_scroll.set)
        self.feat_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        feat_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.feat_tree.tag_configure("phase_header", background="#bbdefb", font=("Segoe UI", 9, "bold"), foreground="#1565c0")
        self.feat_tree.tag_configure("group_header", background="#e3f2fd", font=("Segoe UI", 9, "bold"))

        # Click on info column → show popup
        self.feat_tree.bind("<ButtonRelease-1>", self._on_feat_tree_click)

        # Tab 2: Comparison table (all images)
        comp_frame = ttk.Frame(right_notebook)
        right_notebook.add(comp_frame, text="Comparison")
        self.comp_tree_frame = comp_frame

        # Tab 3: Processing summary
        summary_frame = ttk.Frame(right_notebook)
        right_notebook.add(summary_frame, text="Summary")

        self.summary_tree = ttk.Treeview(
            summary_frame,
            columns=("image", "features", "time"),
            show="headings",
            selectmode="browse",
        )
        self.summary_tree.heading("image", text="Image")
        self.summary_tree.heading("features", text="Features")
        self.summary_tree.heading("time", text="Time(s)")
        self.summary_tree.column("image", width=220)
        self.summary_tree.column("features", width=70, anchor="center")
        self.summary_tree.column("time", width=70, anchor="e")

        sum_scroll = ttk.Scrollbar(summary_frame, orient=tk.VERTICAL, command=self.summary_tree.yview)
        self.summary_tree.config(yscrollcommand=sum_scroll.set)
        self.summary_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sum_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def _bind_events(self):
        self.image_listbox.bind("<<ListboxSelect>>", self._on_image_select)
        self.bind("<Control-o>", lambda e: self._open_files())
        self.bind("<Control-Shift-O>", lambda e: self._open_folder())

    # ── File operations ──
    def _open_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Images",
            filetypes=[("BMP files", "*.bmp"), ("All images", "*.bmp;*.png;*.tif;*.jpg"), ("All", "*.*")],
        )
        if paths:
            for p in paths:
                if p not in self.image_paths:
                    self.image_paths.append(p)
                    self.image_listbox.insert(tk.END, Path(p).name)

    def _open_folder(self):
        folder = filedialog.askdirectory(title="Select Image Folder")
        if folder:
            folder = Path(folder)
            for p in sorted(folder.glob("*.bmp")):
                sp = str(p)
                if sp not in self.image_paths:
                    self.image_paths.append(sp)
                    self.image_listbox.insert(tk.END, p.name)

    def _clear_all(self):
        self.image_paths.clear()
        self.image_listbox.delete(0, tk.END)
        self.feature_values.clear()
        self.phase_images.clear()
        self.proc_times.clear()
        self.current_index = -1
        for tab_id, canvas in self.phase_canvases.items():
            canvas.delete("all")
            self.phase_photo_refs[tab_id] = None
        for item in self.feat_tree.get_children():
            self.feat_tree.delete(item)
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        for widget in self.comp_tree_frame.winfo_children():
            widget.destroy()
        self.status_label.config(text="Ready", foreground="gray")

    # ── Analysis ──
    def _run_analysis(self):
        if not self.image_paths:
            messagebox.showwarning("Warning", "No images loaded.")
            return
        if self.processing:
            return

        self.processing = True
        self.run_btn.config(state="disabled")
        self.progress["maximum"] = len(self.image_paths)
        self.progress["value"] = 0

        thread = threading.Thread(target=self._analysis_worker, daemon=True)
        thread.start()

    def _analysis_worker(self):
        from core.feature_engine import FeatureEngine
        from config.camera import CameraConfig
        from gui.phase_renderer import render_all_phases
        import time

        engine = FeatureEngine(CameraConfig())

        for i, img_path in enumerate(self.image_paths):
            name = Path(img_path).name
            self.after(0, lambda n=name, idx=i: self._update_status(
                f"Processing {n}... ({idx+1}/{len(self.image_paths)})", "orange"))

            t0 = time.time()

            # Extract features
            features = engine.extract(img_path)
            self.feature_values[img_path] = {fid: fr.value for fid, fr in features.items()}

            # Generate phase images
            phase_imgs = render_all_phases(img_path, CameraConfig())
            self.phase_images[img_path] = phase_imgs

            elapsed = time.time() - t0
            self.proc_times[img_path] = round(elapsed, 2)

            self.after(0, lambda idx=i: self.progress.configure(value=idx + 1))

        self.after(0, self._analysis_complete)

    def _analysis_complete(self):
        self.processing = False
        self.run_btn.config(state="normal")
        self.status_label.config(
            text=f"Done - {len(self.image_paths)} images processed", foreground="green")

        self._update_summary()
        self._update_comparison()

        if self.image_paths:
            self.image_listbox.selection_clear(0, tk.END)
            self.image_listbox.selection_set(0)
            self._on_image_select(None)

    def _update_status(self, text, color):
        self.status_label.config(text=text, foreground=color)

    # ── Display ──
    def _on_image_select(self, event):
        sel = self.image_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx == self.current_index:
            return
        self.current_index = idx
        img_path = self.image_paths[idx]

        if img_path in self.phase_images:
            self._display_phases(img_path)
        if img_path in self.feature_values:
            self._display_features(img_path)

    def _display_phases(self, img_path):
        from PIL import Image, ImageTk

        phases = self.phase_images.get(img_path, {})

        ref_canvas = self.phase_canvases.get("Original")
        ref_canvas.update_idletasks()
        ref_cw = max(ref_canvas.winfo_width(), 400)
        ref_ch = max(ref_canvas.winfo_height(), 400)

        for tab_id, canvas in self.phase_canvases.items():
            canvas.delete("all")
            self.phase_photo_refs[tab_id] = None

            pil_img = phases.get(tab_id)
            if pil_img is None:
                continue

            iw, ih = pil_img.size
            scale = min(ref_cw / iw, ref_ch / ih)
            new_w = int(iw * scale)
            new_h = int(ih * scale)

            resized = pil_img.resize((new_w, new_h), Image.LANCZOS)
            photo = ImageTk.PhotoImage(resized)

            canvas.create_image(ref_cw // 2, ref_ch // 2, image=photo, anchor="center")
            self.phase_photo_refs[tab_id] = photo

    def _display_features(self, img_path):
        for item in self.feat_tree.get_children():
            self.feat_tree.delete(item)

        values = self.feature_values.get(img_path, {})
        from config.feature_registry import FEATURE_REGISTRY

        current_phase = ""
        current_cat = ""
        for fspec in FEATURE_REGISTRY:
            # Phase header (top-level grouping by quality classification priority)
            phase_label = fspec.quality_phase.value if hasattr(fspec, 'quality_phase') else ""
            if phase_label and phase_label != current_phase:
                current_phase = phase_label
                self.feat_tree.insert("", "end", values=(
                    f"═══ {current_phase} ═══", "", ""
                ), tags=("phase_header",))
                current_cat = ""  # reset category on phase change

            # Category sub-header
            if fspec.category != current_cat:
                current_cat = fspec.category
                self.feat_tree.insert("", "end", values=(
                    f"── {current_cat} ({fspec.group.value}) ──", "", ""
                ), tags=("group_header",))

            val = values.get(fspec.id)
            val_str = f"{val:.6f}" if val is not None else "N/A"
            self.feat_tree.insert("", "end", iid=fspec.id, values=(fspec.id, val_str, "[i]"))

    def _on_feat_tree_click(self, event):
        """Handle click on feature tree — show info popup if [i] column clicked."""
        region = self.feat_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.feat_tree.identify_column(event.x)
        item_id = self.feat_tree.identify_row(event.y)
        if not item_id or col != "#3":  # #3 = info column
            return

        from config.feature_registry import get_feature
        try:
            fspec = get_feature(item_id)
        except KeyError:
            return

        self._show_feature_info(fspec, event)

    def _show_feature_info(self, fspec, event):
        """Show popup window with feature details."""
        popup = tk.Toplevel(self)
        popup.title(f"Feature Info: {fspec.name}")
        popup.geometry("560x580")
        popup.resizable(False, False)
        popup.transient(self)
        popup.grab_set()

        # Position near click
        popup.geometry(f"+{event.x_root + 10}+{event.y_root - 50}")

        # Scrollable content
        outer_frame = ttk.Frame(popup)
        outer_frame.pack(fill=tk.BOTH, expand=True)

        canvas_popup = tk.Canvas(outer_frame)
        scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas_popup.yview)
        main_frame = ttk.Frame(canvas_popup, padding=15)

        main_frame.bind("<Configure>", lambda e: canvas_popup.configure(scrollregion=canvas_popup.bbox("all")))
        canvas_popup.create_window((0, 0), window=main_frame, anchor="nw")
        canvas_popup.configure(yscrollcommand=scrollbar.set)

        canvas_popup.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas_popup.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas_popup.bind_all("<MouseWheel>", _on_mousewheel)
        popup.protocol("WM_DELETE_WINDOW", lambda: (canvas_popup.unbind_all("<MouseWheel>"), popup.destroy()))

        # Title
        ttk.Label(main_frame, text=fspec.name,
                  font=("Segoe UI", 14, "bold")).pack(anchor="w")

        # Phase badge
        phase_label = fspec.quality_phase.value if hasattr(fspec, 'quality_phase') else ""
        subtitle = f"ID: {fspec.id}  |  Group {fspec.group.value}  |  {fspec.category}"
        if phase_label:
            subtitle += f"\n{phase_label}"
        ttk.Label(main_frame, text=subtitle,
                  foreground="gray").pack(anchor="w", pady=(0, 10))

        ttk.Separator(main_frame, orient="horizontal").pack(fill="x", pady=5)

        # Method
        ttk.Label(main_frame, text="[Extraction Method]",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(5, 2))
        method_text = tk.Text(main_frame, height=3, wrap="word", font=("Segoe UI", 9),
                              relief="flat", bg="#f5f5f5", padx=8, pady=5)
        method_text.insert("1.0", fspec.method or "N/A")
        method_text.config(state="disabled")
        method_text.pack(fill="x")

        # Physical meaning
        ttk.Label(main_frame, text="[Physical Meaning]",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(10, 2))
        meaning_text = tk.Text(main_frame, height=3, wrap="word", font=("Segoe UI", 9),
                               relief="flat", bg="#f5f5f5", padx=8, pady=5)
        meaning_text.insert("1.0", fspec.physical_meaning or "N/A")
        meaning_text.config(state="disabled")
        meaning_text.pack(fill="x")

        # Quality note
        ttk.Label(main_frame, text="[Quality Note]",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(10, 2))
        quality_text = tk.Text(main_frame, height=3, wrap="word", font=("Segoe UI", 9),
                               relief="flat", bg="#fff8e1", padx=8, pady=5)
        quality_text.insert("1.0", fspec.quality_note or "N/A")
        quality_text.config(state="disabled")
        quality_text.pack(fill="x")

        # Classification Guide (new section)
        guide_content = getattr(fspec, 'classification_guide', '') or ''
        if guide_content:
            ttk.Label(main_frame, text="[Classification Guide — 정상/불량 분류 가이드]",
                      font=("Segoe UI", 10, "bold"), foreground="#1565c0").pack(anchor="w", pady=(10, 2))
            guide_text = tk.Text(main_frame, height=5, wrap="word", font=("Segoe UI", 9),
                                 relief="flat", bg="#e8f5e9", padx=8, pady=5)
            guide_text.insert("1.0", guide_content)
            guide_text.config(state="disabled")
            guide_text.pack(fill="x")

        # Close button
        ttk.Button(main_frame, text="Close",
                   command=lambda: (canvas_popup.unbind_all("<MouseWheel>"), popup.destroy())
                   ).pack(pady=(15, 0))

    def _show_phase_info(self, tab_id):
        """Show popup explaining the image-processing logic of a visualization tab."""
        from gui.phase_renderer import PHASE_DESCRIPTIONS

        desc = PHASE_DESCRIPTIONS.get(tab_id)
        if not desc:
            return

        popup = tk.Toplevel(self)
        popup.title(f"처리 설명: {desc['title']}")
        popup.geometry("620x600")
        popup.transient(self)
        popup.grab_set()

        # Scrollable content
        outer = ttk.Frame(popup)
        outer.pack(fill=tk.BOTH, expand=True)
        canvas_popup = tk.Canvas(outer)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas_popup.yview)
        body = ttk.Frame(canvas_popup, padding=15)

        body.bind("<Configure>", lambda e: canvas_popup.configure(scrollregion=canvas_popup.bbox("all")))
        canvas_popup.create_window((0, 0), window=body, anchor="nw")
        canvas_popup.configure(yscrollcommand=scrollbar.set)
        canvas_popup.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_mousewheel(event):
            canvas_popup.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas_popup.bind_all("<MouseWheel>", _on_mousewheel)
        popup.protocol("WM_DELETE_WINDOW",
                       lambda: (canvas_popup.unbind_all("<MouseWheel>"), popup.destroy()))

        # Title
        ttk.Label(body, text=desc["title"], font=("Segoe UI", 14, "bold")).pack(anchor="w")
        ttk.Label(body, text=f"탭 ID: {tab_id}", foreground="gray").pack(anchor="w", pady=(0, 8))
        ttk.Separator(body, orient="horizontal").pack(fill="x", pady=5)

        sections = [
            ("[목적 — 무엇을 보는 화면인가]", desc.get("purpose", ""), "#f5f5f5"),
            ("[처리 로직 — 어떤 이미지 처리를 거치는가]", desc.get("method", ""), "#e3f2fd"),
            ("[해석 — 어떻게 읽는가]", desc.get("interpret", ""), "#fff8e1"),
            ("[관련 Feature]", desc.get("related", ""), "#e8f5e9"),
        ]
        for label, content, bg in sections:
            ttk.Label(body, text=label, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(10, 2))
            txt = tk.Text(body, height=4, wrap="word", font=("Segoe UI", 9),
                          relief="flat", bg=bg, padx=8, pady=5)
            txt.insert("1.0", content or "N/A")
            txt.config(state="disabled")
            txt.pack(fill="x")

        ttk.Button(body, text="Close",
                   command=lambda: (canvas_popup.unbind_all("<MouseWheel>"), popup.destroy())
                   ).pack(pady=(15, 0))

    def _show_capture_conditions(self):
        """Show the reference camera / microscope capture conditions."""
        from config.camera import CAPTURE_EQUIPMENT, CAPTURE_CONDITIONS, CAPTURE_NOTES

        popup = tk.Toplevel(self)
        popup.title("촬영 조건 (Capture Conditions Reference)")
        popup.geometry("660x600")
        popup.transient(self)
        popup.grab_set()

        body = ttk.Frame(popup, padding=12)
        body.pack(fill=tk.BOTH, expand=True)

        ttk.Label(body, text="촬영 조건 Reference", font=("Segoe UI", 14, "bold")).pack(anchor="w")
        ttk.Label(body, text=f"장비: {CAPTURE_EQUIPMENT}", foreground="gray").pack(anchor="w", pady=(0, 6))
        ttk.Label(body, text="아래 조건을 동일하게 유지하여 촬영하세요.",
                  foreground="#1565c0").pack(anchor="w", pady=(0, 8))

        # Settings table
        cols = ("item", "value", "note")
        tree = ttk.Treeview(body, columns=cols, show="headings",
                            height=len(CAPTURE_CONDITIONS), selectmode="none")
        tree.heading("item", text="항목")
        tree.heading("value", text="설정값")
        tree.heading("note", text="비고")
        tree.column("item", width=150)
        tree.column("value", width=140, anchor="center")
        tree.column("note", width=300)
        tree.tag_configure("odd", background="#f5f5f5")
        for i, (item, value, note) in enumerate(CAPTURE_CONDITIONS):
            tree.insert("", "end", values=(item, value, note),
                        tags=("odd",) if i % 2 else ())
        tree.pack(fill=tk.X, pady=(0, 10))

        # Notes
        ttk.Label(body, text="[유의사항]", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        notes = tk.Text(body, height=9, wrap="word", font=("Segoe UI", 9),
                        relief="flat", bg="#fff8e1", padx=8, pady=6)
        notes.insert("1.0", CAPTURE_NOTES)
        notes.config(state="disabled")
        notes.pack(fill=tk.BOTH, expand=True)

        ttk.Button(body, text="Close", command=popup.destroy).pack(pady=(10, 0))

    # ── Tab image copy / download ──
    def _current_phase_image(self, tab_id):
        """Return (img_path, PIL.Image) for the given tab of the selected image, or (None, None)."""
        if self.current_index < 0 or self.current_index >= len(self.image_paths):
            messagebox.showwarning("Warning", "먼저 좌측에서 이미지를 선택하세요.")
            return None, None
        img_path = self.image_paths[self.current_index]
        phases = self.phase_images.get(img_path)
        if not phases or phases.get(tab_id) is None:
            messagebox.showwarning(
                "Warning", "이 탭의 이미지가 아직 없습니다.\n먼저 '▶ Extract Features'를 실행하세요.")
            return None, None
        return img_path, phases[tab_id]

    def _copy_phase_image(self, tab_id):
        """Copy the current tab image to the clipboard (pasteable into Excel/PPT)."""
        img_path, pil_img = self._current_phase_image(tab_id)
        if pil_img is None:
            return
        try:
            _image_to_clipboard(pil_img)
        except Exception as e:
            messagebox.showerror("복사 실패", f"클립보드 복사에 실패했습니다:\n{e}")
            return
        self._update_status(f"복사됨: {tab_id} — Excel/PPT에 붙여넣기(Ctrl+V)", "blue")

    def _download_phase_image(self, tab_id):
        """Save the current tab image to disk without lossy compression."""
        img_path, pil_img = self._current_phase_image(tab_id)
        if pil_img is None:
            return
        default_name = f"{Path(img_path).stem}_{tab_id}.png"
        out_dir = Path(__file__).resolve().parent.parent / "output"
        path = filedialog.asksaveasfilename(
            title="이미지 저장 (무압축)",
            defaultextension=".png",
            initialfile=default_name,
            initialdir=str(out_dir),
            filetypes=[
                ("PNG (무손실)", "*.png"),
                ("BMP (무압축)", "*.bmp"),
                ("TIFF (무압축)", "*.tif"),
                ("All", "*.*"),
            ],
        )
        if not path:
            return
        try:
            _save_image_lossless(pil_img, path)
        except Exception as e:
            messagebox.showerror("저장 실패", f"이미지 저장에 실패했습니다:\n{e}")
            return
        self._update_status(f"저장됨: {Path(path).name}", "green")

    def _update_summary(self):
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)

        for img_path in self.image_paths:
            name = Path(img_path).name
            n_feat = len(self.feature_values.get(img_path, {}))
            elapsed = self.proc_times.get(img_path, "-")
            self.summary_tree.insert("", "end", values=(name, n_feat, elapsed))

    def _update_comparison(self):
        for widget in self.comp_tree_frame.winfo_children():
            widget.destroy()

        if not self.feature_values:
            return

        from config.feature_registry import FEATURE_REGISTRY

        img_names = [Path(p).stem[-20:] for p in self.image_paths]
        columns = ("feature",) + tuple(f"img_{i}" for i in range(len(self.image_paths)))

        tree = ttk.Treeview(self.comp_tree_frame, columns=columns, show="headings", selectmode="browse")
        tree.heading("feature", text="Feature")
        tree.column("feature", width=140)

        for i, name in enumerate(img_names):
            col_id = f"img_{i}"
            tree.heading(col_id, text=name)
            tree.column(col_id, width=100, anchor="e")

        tree.tag_configure("group_header", background="#e3f2fd", font=("Segoe UI", 9, "bold"))

        current_cat = ""
        for fspec in FEATURE_REGISTRY:
            if fspec.category != current_cat:
                current_cat = fspec.category
                row_vals = [f"── {current_cat} ──"] + [""] * len(self.image_paths)
                tree.insert("", "end", values=tuple(row_vals), tags=("group_header",))

            row_vals = [fspec.id]
            for img_path in self.image_paths:
                vals = self.feature_values.get(img_path, {})
                v = vals.get(fspec.id)
                row_vals.append(f"{v:.4f}" if v is not None else "")
            tree.insert("", "end", values=tuple(row_vals))

        vsb = ttk.Scrollbar(self.comp_tree_frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(self.comp_tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.comp_tree_frame.grid_rowconfigure(0, weight=1)
        self.comp_tree_frame.grid_columnconfigure(0, weight=1)

    # ── Export ──
    def _export_csv(self):
        if not self.feature_values:
            messagebox.showwarning("Warning", "No data to export. Run analysis first.")
            return

        path = filedialog.asksaveasfilename(
            title="Export CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialdir=str(Path(__file__).resolve().parent.parent / "output"),
        )
        if not path:
            return

        from config.feature_registry import FEATURE_REGISTRY

        headers = ["image"] + [f.id for f in FEATURE_REGISTRY]
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for img_path in self.image_paths:
                vals = self.feature_values.get(img_path, {})
                row = [Path(img_path).name]
                for fspec in FEATURE_REGISTRY:
                    v = vals.get(fspec.id)
                    row.append(f"{v:.6f}" if v is not None else "")
                writer.writerow(row)

        messagebox.showinfo("Export", f"CSV saved: {path}")

    def _export_excel(self):
        if not self.feature_values:
            messagebox.showwarning("Warning", "No data to export. Run analysis first.")
            return

        path = filedialog.asksaveasfilename(
            title="Export Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=str(Path(__file__).resolve().parent.parent / "output"),
        )
        if not path:
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from config.feature_registry import FEATURE_REGISTRY

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feature Values"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        # Headers: Feature ID | Category | Group | Image1 | Image2 | ...
        headers = ["Feature", "Category", "Group"]
        img_names = [Path(p).name for p in self.image_paths]
        headers += img_names

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        row_idx = 2
        current_cat = ""
        for fspec in FEATURE_REGISTRY:
            # Category separator row
            if fspec.category != current_cat:
                current_cat = fspec.category
                cell = ws.cell(row=row_idx, column=1, value=f"{current_cat} ({fspec.group.value})")
                cell.fill = cat_fill
                cell.font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).fill = cat_fill
                    ws.cell(row=row_idx, column=c).border = thin_border
                row_idx += 1

            ws.cell(row=row_idx, column=1, value=fspec.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=fspec.category).border = thin_border
            ws.cell(row=row_idx, column=3, value=fspec.group.value).border = thin_border

            for img_i, img_path in enumerate(self.image_paths):
                vals = self.feature_values.get(img_path, {})
                v = vals.get(fspec.id)
                cell = ws.cell(row=row_idx, column=4 + img_i, value=round(v, 6) if v is not None else None)
                cell.border = thin_border
                cell.number_format = "0.000000"
            row_idx += 1

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(path)
        messagebox.showinfo("Export", f"Excel saved: {path}")

    def _export_all_images(self):
        """Save every phase image (Original + all processed) for selected source images.

        - Multiple images loaded → user picks which ones via a selection dialog.
        - Single image → export it directly without a dialog.
        Each image's phases are saved as lossless PNG: {stem}_{tab_id}.png
        """
        if not self.phase_images:
            messagebox.showwarning(
                "Warning", "다운로드할 이미지가 없습니다.\n먼저 '▶ Extract Features'를 실행하세요.")
            return

        # Source images that actually have rendered phases
        available = [p for p in self.image_paths
                     if self.phase_images.get(p)]
        if not available:
            messagebox.showwarning(
                "Warning", "추출된 이미지가 없습니다.\n먼저 '▶ Extract Features'를 실행하세요.")
            return

        # Single image → no selection needed; multiple → let the user choose
        if len(available) == 1:
            selected = available
        else:
            selected = self._select_images_dialog(available)
            if not selected:  # cancelled or nothing chosen
                return

        out_dir = filedialog.askdirectory(
            title="이미지를 저장할 폴더 선택",
            initialdir=str(Path(__file__).resolve().parent.parent / "output"),
        )
        if not out_dir:
            return
        out_dir = Path(out_dir)

        saved = 0
        failed = []
        for img_path in selected:
            stem = Path(img_path).stem
            for tab_id, pil_img in self.phase_images.get(img_path, {}).items():
                if pil_img is None:
                    continue
                try:
                    _save_image_lossless(pil_img, str(out_dir / f"{stem}_{tab_id}.png"))
                    saved += 1
                except Exception as e:
                    failed.append(f"{stem}_{tab_id}: {e}")

        msg = (f"이미지 {len(selected)}개 × 처리 단계별 = 총 {saved}개 파일 저장 완료.\n"
               f"형식: PNG(무압축)\n위치: {out_dir}")
        if failed:
            msg += f"\n\n실패 {len(failed)}건:\n" + "\n".join(failed[:5])
        messagebox.showinfo("다운로드 완료", msg)
        self._update_status(f"{saved}개 이미지 파일 저장됨 → {out_dir.name}", "green")

    def _select_images_dialog(self, available):
        """Modal dialog to pick which source images to export. Returns list of paths (or None)."""
        dlg = tk.Toplevel(self)
        dlg.title("다운로드할 이미지 선택")
        dlg.geometry("440x480")
        dlg.transient(self)
        dlg.grab_set()

        ttk.Label(dlg, text="다운로드할 이미지를 선택하세요 (Ctrl/Shift로 다중 선택)",
                  padding=8).pack(anchor="w")

        frame = ttk.Frame(dlg)
        frame.pack(fill=tk.BOTH, expand=True, padx=8)
        lb = tk.Listbox(frame, selectmode=tk.EXTENDED, font=("Segoe UI", 9))
        sb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=lb.yview)
        lb.config(yscrollcommand=sb.set)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        for p in available:
            lb.insert(tk.END, Path(p).name)

        # Pre-select whatever is highlighted in the main list; else select all
        main_sel = set(self.image_listbox.curselection())
        if main_sel:
            for i, p in enumerate(available):
                if self.image_paths.index(p) in main_sel:
                    lb.selection_set(i)
        else:
            lb.selection_set(0, tk.END)

        result = {"paths": None}

        def _ok():
            result["paths"] = [available[i] for i in lb.curselection()]
            dlg.destroy()

        btns = ttk.Frame(dlg)
        btns.pack(fill=tk.X, padx=8, pady=8)
        ttk.Button(btns, text="전체 선택",
                   command=lambda: lb.selection_set(0, tk.END)).pack(side=tk.LEFT)
        ttk.Button(btns, text="취소", command=dlg.destroy).pack(side=tk.RIGHT, padx=(4, 0))
        ttk.Button(btns, text="다운로드", command=_ok).pack(side=tk.RIGHT)

        dlg.wait_window()
        return result["paths"]


def main():
    app = InspectionApp()
    app.mainloop()


if __name__ == "__main__":
    main()
